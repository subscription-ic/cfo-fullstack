"""
Upload questions from Excel files in upload_questions/ folder to the
Supabase `predicted_qa` table.

Table schema:
  id               uuid (auto-generated)
  company          text
  predicted_question text
  suggested_answer text
  category         text
  risk             text
  created_at       timestamptz (auto-generated)

Usage:
  python upload_questions_to_supabase.py
"""

import os
import uuid
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from supabase import create_client, Client


# ── Auto-discover .env by walking up directory tree ───────────────────────────
def find_env_file(start: Path) -> Path | None:
    """Search start dir and every parent for a .env file."""
    for directory in [start, *start.parents]:
        candidate = directory / ".env"
        if candidate.is_file():
            return candidate
    return None

env_path = find_env_file(Path(__file__).parent)
if env_path is None:
    raise FileNotFoundError("Could not find a .env file in this folder or any parent folder.")

print(f"Using .env at: {env_path}")
load_dotenv(dotenv_path=env_path)

SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY")

if not SUPABASE_URL or not SUPABASE_KEY:
    raise ValueError("SUPABASE_URL and SUPABASE_KEY must be set in the .env file.")

# ── Supabase client ───────────────────────────────────────────────────────────
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ── Expected Excel column → Supabase column mapping ──────────────────────────
# Keys   = exact header names in the Excel file (case-insensitive comparison)
# Values = column names in the `predicted_qa` table
COLUMN_MAP = {
    "company":            "company",
    "predicted_question": "predicted_question",
    "suggested_answer":   "suggested_answer",
    "category":           "category",
    "risk":               "risk",
}

UPLOAD_DIR = Path(__file__).parent / "upload_questions"
TABLE_NAME = "predicted_qa"
BATCH_SIZE = 50          # rows per upsert call


def read_excel(filepath: Path) -> list[dict]:
    """Read an Excel file and return a list of row dicts ready for Supabase."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    raw_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    # Normalise header names (strip whitespace, lower-case)
    headers = [str(h).strip().lower() if h is not None else "" for h in raw_headers]

    # Build index → supabase column name
    col_indices: dict[int, str] = {}
    for idx, h in enumerate(headers):
        if h in COLUMN_MAP:
            col_indices[idx] = COLUMN_MAP[h]

    if not col_indices:
        raise ValueError(
            f"No recognised columns found in {filepath.name}. "
            f"Expected one of: {list(COLUMN_MAP.keys())}. Got: {headers}"
        )

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record: dict = {"id": str(uuid.uuid4())}
        for idx, supabase_col in col_indices.items():
            val = row[idx] if idx < len(row) else None
            record[supabase_col] = str(val).strip() if val is not None else None

        # Skip fully-empty rows
        values = [v for k, v in record.items() if k != "id"]
        if all(v is None or v == "" for v in values):
            continue

        rows.append(record)

    wb.close()
    return rows


def upload_rows(rows: list[dict], source_file: str) -> None:
    """Insert rows into Supabase in batches."""
    total = len(rows)
    uploaded = 0

    for start in range(0, total, BATCH_SIZE):
        batch = rows[start : start + BATCH_SIZE]
        response = supabase.table(TABLE_NAME).insert(batch).execute()

        # supabase-py v2 raises on error automatically; log success count
        uploaded += len(batch)
        print(f"  [{source_file}] Uploaded {uploaded}/{total} rows …")

    print(f"  [{source_file}] ✅ Done — {total} rows inserted.")


def main() -> None:
    excel_files = sorted(UPLOAD_DIR.glob("*.xlsx")) + sorted(UPLOAD_DIR.glob("*.xls"))

    if not excel_files:
        print(f"No Excel files found in {UPLOAD_DIR}")
        return

    print(f"Found {len(excel_files)} file(s) in {UPLOAD_DIR}\n")

    for filepath in excel_files:
        print(f"Processing: {filepath.name}")
        try:
            rows = read_excel(filepath)
            print(f"  Parsed {len(rows)} data rows.")
            if rows:
                upload_rows(rows, filepath.name)
        except Exception as exc:
            print(f"  ❌ Error processing {filepath.name}: {exc}")

    print("\nAll files processed.")


if __name__ == "__main__":
    main()
