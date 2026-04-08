import openpyxl
import uuid

wb = openpyxl.load_workbook('upload_questions/hdfc.xlsx', data_only=True)
ws = wb.active

headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]

def esc(v):
    if v is None:
        return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"

vals = []
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    # skip blank rows
    if all((d.get(k) is None or str(d.get(k)).strip() == '') for k in ['company','predicted_question','suggested_answer','category','risk']):
        continue
    row_id = str(uuid.uuid4())
    vals.append(
        f"  ('{row_id}', {esc(d.get('company'))}, {esc(d.get('predicted_question'))}, "
        f"{esc(d.get('suggested_answer'))}, {esc(d.get('category'))}, {esc(d.get('risk'))})"
    )

sql = (
    "INSERT INTO predicted_qa (id, company, predicted_question, suggested_answer, category, risk) VALUES\n"
    + ",\n".join(vals)
    + ";"
)

with open('insert_predicted_qa.sql', 'w', encoding='utf-8') as f:
    f.write(sql)

print(f"Done. {len(vals)} rows written to insert_predicted_qa.sql")
